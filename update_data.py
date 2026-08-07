"""
영등포 매출 데이터 업데이트 스크립트
실행: python update_data.py
"""
import json
import os
import re
from datetime import datetime
from openpyxl import load_workbook
try:
    from playwright.sync_api import sync_playwright
    _HAS_PLAYWRIGHT = True
except ImportError:
    _HAS_PLAYWRIGHT = False

NAVER_PLACE_ID = "1708007751"

if os.name == "nt":  # Windows
    _ONEDRIVE = r"C:\Users\안태환\Desktop\OneDrive"
else:  # macOS
    _ONEDRIVE = os.path.expanduser("~/Library/CloudStorage/OneDrive-개인")

_BASE      = os.path.join(_ONEDRIVE, "01. 외식업 신규 ( 26.08~ )", "00. 영등포 (26.08 ~ )", "00. 매출 & 손익")
EXCEL_PATH = os.path.join(_BASE, "01. 매출", "H.영등포 월간 매출 _ 26Y.xlsx")
PNL_PATH   = os.path.join(_BASE, "00. 실적", "H.영등포 누적 실적 _ 25Y~.xlsx")
PMIX_PATH  = os.path.join(_BASE, "02.P-MIX", "H.영등포 P-MIX _ 26Y.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "data", "data.json")

WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]

def parse_sheet(ws, month):
    days = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        date_25y = row[1]  # col B: 25Y 날짜
        if not isinstance(date_25y, datetime):
            continue

        # 26Y 날짜 = 25Y 날짜 + 1년
        try:
            date_26y = date_25y.replace(year=date_25y.year + 1)
        except ValueError:
            # 2월 29일 윤년 처리
            date_26y = date_25y.replace(year=date_25y.year + 1, day=28)

        target = row[16] or 0       # col Q: 26Y 목표
        sales_l = row[18] or 0      # col S: 26Y 홀 매출
        sales_d = row[19] or 0      # col T: 26Y 배달 매출
        count_l = row[21] or 0      # col V: 26Y 홀 객수
        count_d = row[22] or 0      # col W: 26Y 배달 객수

        actual = sales_l + sales_d
        count = count_l + count_d
        avg_spend = round(actual / count, 1) if count > 0 else 0

        # 전년 데이터
        prev_target = row[4] or 0   # col E: 25Y 목표
        prev_actual = row[5] or 0   # col F: 25Y 실적
        prev_count = (row[8] or 0)  # col I: 25Y 객수

        # 실적이 없고 목표도 없는 행은 빈 행으로 간주
        if target == 0 and sales_l == 0 and sales_d == 0:
            continue

        days.append({
            "date": date_26y.strftime("%Y-%m-%d"),
            "weekday": WEEKDAY_KO[date_26y.weekday()],
            "target": int(target),
            "actual": int(actual),
            "sales_l": int(sales_l),
            "sales_d": int(sales_d),
            "count": int(count),
            "count_l": int(count_l),
            "count_d": int(count_d),
            "avg_spend": avg_spend,
            "prev_target": int(prev_target),
            "prev_actual": int(prev_actual),
            "prev_count": int(prev_count),
        })
    return days

def parse_pnl():
    wb = load_workbook(PNL_PATH, data_only=True)
    ws = wb['26Y 실적']
    rows = list(ws.iter_rows(min_row=1, max_row=220, max_col=20, values_only=True))

    idx = {}
    sub_labels = {
        '재료비': ['1) 식료재료비', '2) 음료재료비', '3) 기타재료비'],
        '노무비': ['1) 고정비', '2) 변동비', '3) 연차수당', '4) 상여금', '5) 퇴직급여'],
        '일반관리비': [
            '1) 복리후생비', '2) 교육훈련비', '3) 여비교통비', '4) 통신비',
            '5) 수도광열비', '6) 세금과공과', '7) 지급수수료', '8) 임차료',
            '9) 감가상각비', '10) 수선비', '11) 소모품비', '12) 도서인쇄비',
            '13) 보험료', '14) 광고선전비', '15) 기타', '16) 라이선스 수수료',
        ],
    }
    sub_idx = {k: {} for k in sub_labels}
    for i, row in enumerate(rows):
        b, c = row[1], row[2]
        if b == '매출': idx['매출'] = i
        elif b == '재료비': idx['재료비'] = i
        elif b == '노무비': idx['노무비'] = i
        elif b == '일반관리비': idx['일반관리비'] = i
        elif b == '영업이익(매장)': idx['영업이익'] = i
        if c:
            for parent, labels in sub_labels.items():
                if c in labels:
                    sub_idx[parent][c] = i

    pnl = {}
    for m in range(1, 13):
        vc = 4 + (m - 1) * 2
        rc = vc + 1

        def v(i):
            row = rows[i]
            val = row[vc] if vc < len(row) else None
            return int(val) if isinstance(val, (int, float)) else 0

        def r(i):
            row = rows[i]
            val = row[rc] if rc < len(row) else None
            return round(float(val) * 100, 1) if isinstance(val, (int, float)) else 0.0

        if '매출' not in idx: continue
        revenue = rows[idx['매출']][vc] if vc < len(rows[idx['매출']]) else None
        if not revenue: continue

        def sub_detail(parent):
            parent_amt = v(idx[parent]) if parent in idx else 0
            result = []
            for lbl in sub_labels[parent]:
                if lbl not in sub_idx[parent]: continue
                amt = v(sub_idx[parent][lbl])
                pct = round(amt / parent_amt * 100, 1) if parent_amt else 0.0
                result.append({"label": lbl.split(') ', 1)[1], "amount": amt, "pct": pct})
            return result

        pnl[str(m)] = {
            "revenue": int(revenue),
            "food_cost": v(idx['재료비']) if '재료비' in idx else 0,
            "food_cost_pct": r(idx['재료비']) if '재료비' in idx else 0.0,
            "food_cost_detail": sub_detail('재료비'),
            "labor_cost": v(idx['노무비']) if '노무비' in idx else 0,
            "labor_cost_pct": r(idx['노무비']) if '노무비' in idx else 0.0,
            "labor_cost_detail": sub_detail('노무비'),
            "ga_cost": v(idx['일반관리비']) if '일반관리비' in idx else 0,
            "ga_cost_pct": r(idx['일반관리비']) if '일반관리비' in idx else 0.0,
            "ga_cost_detail": sub_detail('일반관리비'),
            "op_profit": v(idx['영업이익']) if '영업이익' in idx else 0,
            "op_profit_pct": r(idx['영업이익']) if '영업이익' in idx else 0.0,
        }
        print(f"  P&L {m}월: 매출 {revenue:,.0f}원, 영업이익 {pnl[str(m)]['op_profit_pct']}%")

    return pnl

def parse_pmix():
    wb = load_workbook(PMIX_PATH, data_only=True)
    pmix = {}
    for sheet_name in wb.sheetnames:
        try:
            m = int(sheet_name)
        except ValueError:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=22, values_only=True))

        def clean_name(s):
            if not s: return ""
            return str(s).replace("(H.CD)", "").replace("(H.CD", "").strip()

        # ── 왼쪽: 홀 메뉴 (col B~H, index 1~7) ──
        all_menus = []
        total_qty = 0; total_revenue = 0
        EXCLUDE_CATEGORIES = {'추가메뉴', '음료'}
        seen = set()
        current_cat = ''
        hall_cost_amt_map = {}   # 홀메뉴명 → E열 원가금액 (프로모션 원가율 계산에 활용)
        for i in range(3, len(rows)):
            row = rows[i]
            cat_cell = row[1]
            if cat_cell and isinstance(cat_cell, str): current_cat = cat_cell.strip()
            name = row[2]; qty = row[6]; price = row[3]; cost_amt = row[4]
            n = clean_name(name)
            # 원가금액은 음료/추가메뉴 포함 전체 홀 메뉴에서 저장 (프로모션 매핑용)
            if n and isinstance(cost_amt, (int, float)) and cost_amt > 0:
                hall_cost_amt_map[n] = cost_amt
            if current_cat in EXCLUDE_CATEGORIES: continue
            # H열(index7)은 엑셀 수식 D*G(VAT포함)이므로 캐시됐을 때 /1.1 필요.
            # 항상 D/1.1*G 계산을 우선. H는 D 또는 G가 None일 때만 보조.
            if isinstance(price, (int, float)) and isinstance(qty, (int, float)):
                rev = price / 1.1 * qty
            elif isinstance(row[7], (int, float)) and row[7] > 0:
                rev = row[7] / 1.1   # H열 캐시값(VAT포함)을 VAT제외로 변환
            else:
                rev = None
            if not n or not isinstance(qty, (int, float)) or not isinstance(price, (int, float)): continue
            if n in seen: continue
            seen.add(n)
            qty_int = int(qty)
            rev_int = int(rev) if isinstance(rev, (int, float)) else 0
            cost_rate = round(cost_amt / (price / 1.1) * 100, 1) \
                if isinstance(cost_amt, (int, float)) and price > 0 else None
            all_menus.append({"name": n, "qty": qty_int, "revenue": rev_int,
                              "category": current_cat, "cost_rate": cost_rate})
            total_qty += qty_int
            total_revenue += rev_int

        # TOP 10: R열(17)·S열(18)=건수 / U열(20)·V열(21)=매출 (시트 행5-14 → 0-indexed 4-13)
        sales_top10 = []
        revenue_top10 = []
        for i in range(4, 14):
            if i >= len(rows): break
            row = rows[i]
            r_name = row[17] if len(row) > 17 else None
            r_qty  = row[18] if len(row) > 18 else None
            u_name = row[20] if len(row) > 20 else None
            v_rev  = row[21] if len(row) > 21 else None
            if isinstance(r_name, str) and isinstance(r_qty, (int, float)):
                sales_top10.append({"name": clean_name(r_name), "qty": int(r_qty)})
            if isinstance(u_name, str) and isinstance(v_rev, (int, float)):
                revenue_top10.append({"name": clean_name(u_name), "revenue": int(v_rev)})

        # 이론원가: R열(17)=소분류 / S열(18)=순매출 / U열(20)=원가율 (시트 행18-28 → 0-indexed 17-27)
        theory_cost = {}
        theory_net_rev = {}
        for i in range(17, 28):
            if i >= len(rows): break
            row = rows[i]
            cat     = row[17] if len(row) > 17 else None
            net_rev = row[18] if len(row) > 18 else None
            rate    = row[20] if len(row) > 20 else None
            if isinstance(cat, str) and cat.strip() and isinstance(rate, (int, float)):
                k = cat.strip()
                theory_cost[k] = round(rate * 100, 1)
                if isinstance(net_rev, (int, float)):
                    theory_net_rev[k] = round(net_rev)

        # 이론원가 폴백: R~V 수식이 캐시 안 된 경우 all_menus 가중평균으로 계산 (홀 기준, 프로모션 제외)
        if not theory_cost and all_menus:
            cat_rev = {}; cat_gp = {}
            for menu in all_menus:
                c = menu['category']; r = menu['revenue']; cr = menu.get('cost_rate')
                if r > 0 and cr is not None:
                    cat_rev[c] = cat_rev.get(c, 0) + r
                    cat_gp[c]  = cat_gp.get(c, 0) + r * (1 - cr / 100)
            for c in cat_rev:
                if cat_rev[c] > 0:
                    theory_cost[c] = round((1 - cat_gp[c] / cat_rev[c]) * 100, 1)
            total_r = sum(cat_rev.values()); total_g = sum(cat_gp.values())
            if total_r > 0:
                theory_cost['합계'] = round((1 - total_g / total_r) * 100, 1)

        # ── 오른쪽: 배달 & 프로모션 (col J~P, index 9~15) ──
        # 프로모션명 → 홀메뉴명 매핑 (M/N열 수식 캐시 없을 때 홀 E열로 원가율 계산)
        PROMO_TO_HALL = {
            "[화]완탕면":          "완탕면",
            "[주말]크리스피라페치킨": "크리스피라페치킨",
            "[3주년]완탕면":       "완탕면",
            "[3주년]라구짜장도삭면": "라구짜장도삭면",
            "[3주년]마파두부뽀짜이판": "돼지고기&마파두부 뽀짜이판",
            "[3주년]라구짜장뽀짜이판": "라구짜장&계란튀김 뽀짜이판",
        }
        delivery_revenue = 0
        promo_revenue = 0
        promo_menus = []
        current_right_cat = ''
        for row in rows[3:]:
            cat = row[9]
            name_cell = row[10]
            # P열(index15)은 엑셀 수식 L*O(VAT포함)이므로 캐시됐을 때 /1.1 필요.
            # 항상 L/1.1*O 계산을 우선. P는 L 또는 O가 None일 때만 보조.
            price_p = row[11] if len(row) > 11 else None
            qty_p0  = row[14] if len(row) > 14 else None
            if isinstance(price_p, (int, float)) and isinstance(qty_p0, (int, float)):
                rev = price_p / 1.1 * qty_p0
            elif isinstance(row[15], (int, float)) and row[15] > 0:
                rev = row[15] / 1.1   # P열 캐시값(VAT포함)을 VAT제외로 변환
            else:
                rev = None
            if cat and isinstance(cat, str) and cat not in ('소분류',):
                current_right_cat = cat.strip()
            # 이름 없는 행(소계행) 제외
            if not (isinstance(name_cell, str) and name_cell): continue
            if not isinstance(rev, (int, float)) or rev <= 0: continue
            if current_right_cat == '배달':
                delivery_revenue += rev
            elif current_right_cat == '프로모션':
                promo_revenue += rev
                n = clean_name(name_cell)
                qty_p = row[14]  # O열: 총건수
                cost_rate_raw = row[13]  # N열: 원가율 (수식 캐시)
                qty_int = int(qty_p) if isinstance(qty_p, (int, float)) and qty_p > 0 else 0
                if isinstance(cost_rate_raw, (int, float)):
                    cost_rate = round(cost_rate_raw * 100, 1)
                else:
                    # 1차 폴백: M열(원가금액) / L열(단가VAT포함)/1.1
                    cost_amt_p = row[12] if len(row) > 12 else None
                    if isinstance(cost_amt_p, (int, float)) and isinstance(price_p, (int, float)) and price_p > 0:
                        cost_rate = round(cost_amt_p / (price_p / 1.1) * 100, 1)
                    else:
                        # 2차 폴백: 홀 메뉴 E열(원가금액)로 계산 (수식 캐시 없을 때)
                        hall_name = PROMO_TO_HALL.get(n)
                        hall_cost = hall_cost_amt_map.get(hall_name) if hall_name else None
                        if isinstance(hall_cost, (int, float)) and isinstance(price_p, (int, float)) and price_p > 0:
                            cost_rate = round(hall_cost / (price_p / 1.1) * 100, 1)
                        else:
                            cost_rate = None
                if n and qty_int > 0:
                    promo_menus.append({
                        "name": n, "qty": qty_int, "revenue": int(rev),
                        "category": "프로모션", "cost_rate": cost_rate
                    })
        # 키인 오류 보정: [화]완탕면 → [3주년]완탕면으로 합산
        MERGE_MAP = {"[화]완탕면": "[3주년]완탕면"}
        merged = {}
        for pm in promo_menus:
            key = MERGE_MAP.get(pm["name"], pm["name"])
            if key in merged:
                merged[key]["qty"]     += pm["qty"]
                merged[key]["revenue"] += pm["revenue"]
            else:
                merged[key] = dict(pm, name=key)
        promo_menus = list(merged.values())

        all_menus += promo_menus
        total_qty  += sum(m["qty"]     for m in promo_menus)
        total_revenue += sum(m["revenue"] for m in promo_menus)

        # TOP10 폴백: 프로모션 합산 후 계산 (홀+프로모션 포함, 배달 제외)
        if not sales_top10 and all_menus:
            ranked = [m for m in all_menus if m['category'] != '배달']
            sales_top10   = [{"name": m["name"], "qty": m["qty"]}
                             for m in sorted(ranked, key=lambda x: -x['qty'])[:10]]
            revenue_top10 = [{"name": m["name"], "revenue": m["revenue"]}
                             for m in sorted(ranked, key=lambda x: -x['revenue'])[:10]]

        if all_menus:
            pmix[str(m)] = {
                "menus_all": all_menus,
                "sales_top10": sales_top10,
                "revenue_top10": revenue_top10,
                "theory_cost": theory_cost,
                "theory_net_rev": theory_net_rev,
                "total_revenue": total_revenue,
                "total_qty": total_qty,
                "delivery_revenue": int(delivery_revenue),
                "promo_revenue": int(promo_revenue),
            }
            print(f"  P-MIX {m}월: 판매TOP{len(sales_top10)} 매출TOP{len(revenue_top10)} 배달 {int(delivery_revenue):,}원 프로모션 {int(promo_revenue):,}원")
    return pmix

def get_naver_review_count():
    if not _HAS_PLAYWRIGHT:
        print("  playwright 미설치 → 리뷰 수 스킵")
        return None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(
                user_agent='Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
                locale='ko-KR'
            )
            page = ctx.new_page()
            page.goto(
                f'https://m.place.naver.com/restaurant/{NAVER_PLACE_ID}/review/visitor',
                wait_until='networkidle', timeout=20000
            )
            html = page.content()
            browser.close()
        m = re.search(r'"hideProductSelectBox"\s*:\s*true\s*,\s*"total"\s*:\s*(\d+)', html)
        if not m:
            m = re.search(r'"total"\s*:\s*(\d+)\s*,\s*"showRecommendationSort"', html)
        return int(m.group(1)) if m else None
    except Exception as e:
        print(f"  리뷰 수 조회 실패: {e}")
        return None


def main():
    print(f"엑셀 파일 읽는 중...")
    wb = load_workbook(EXCEL_PATH, data_only=True)

    all_data = {}
    for month in range(1, 13):
        sheet_name = str(month)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        days = parse_sheet(ws, month)
        if days:
            all_data[sheet_name] = days
            print(f"  {month}월: {len(days)}일 데이터 읽음")

    print(f"\nP&L 파일 읽는 중...")
    pnl_data = parse_pnl()

    # 월별 고객수·객단가를 P&L에 추가
    for m_str, days in all_data.items():
        if m_str not in pnl_data: continue
        total_count = sum(d['count'] for d in days if d['actual'] > 0)
        total_actual = sum(d['actual'] for d in days)
        avg_spend = round(total_actual / total_count) if total_count else 0
        pnl_data[m_str]['customer_count'] = total_count
        pnl_data[m_str]['avg_spend'] = avg_spend

    print(f"\nP-MIX 파일 읽는 중...")
    pmix_data = parse_pmix()

    print(f"\n네이버 리뷰 수 가져오는 중...")
    review_count = get_naver_review_count()
    today = datetime.now().strftime("%Y-%m-%d")

    # 기존 data.json의 reviews 이력 유지
    reviews_history = {}
    if os.path.exists(OUTPUT_PATH):
        try:
            with open(OUTPUT_PATH, encoding="utf-8") as f:
                old = json.load(f)
            reviews_history = old.get("reviews", {})
        except Exception:
            pass
    if review_count is not None:
        reviews_history[today] = review_count
        print(f"  {today}: {review_count:,}개")
    else:
        print(f"  리뷰 수 가져오기 실패 (이전 이력 유지)")

    output = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "store": "영등포",
        "store_code": "영등포",
        "months": all_data,
        "pnl": pnl_data,
        "pmix": pmix_data,
        "reviews": reviews_history,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n완료! → {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
